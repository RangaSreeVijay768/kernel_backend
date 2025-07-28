from flask import Flask, request, send_file, jsonify
import pandas as pd
import re
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from calculations_nt import generate_pages_for_tag as calculate_values_nt, TagInfo as TagInfoNT, TagDir as TagDirNT
from calculations_aline import calculate_values as calculate_values_aline, TagInfo as TagInfoAline, TagDir as TagDirAline
from calculations_adj import generate_pages_for_tag as calculate_values_adj, TagInfo as TagInfoAdj, TagDir as TagDirAdj

app = Flask(__name__)

def extract_tag_columns(columns):
    tag_cols = []
    for c in columns:
        col_str = str(c).strip()
        if re.match(r"^\d+/[MD]$", col_str):
            tag_cols.append(c)
        else:
            print(f"Column '{col_str}' does not match regex r'^\\d+/[MD]$'")
    return tag_cols

def get_style_elements():
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    font = Font(name="Times New Roman", size=11)
    bold_font = Font(name="Times New Roman", size=11, bold=True)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True, indent=1)
    return border, fill, font, bold_font, align

def format_sheet(ws, df, tag_title):
    from math import ceil

    border, fill, font, bold_font, align = get_style_elements()
    predefined_cols = ["FIELD NAME / DESCRIPTION", "BIT POSITION", "Size (Bits)"]
    tag_columns = [col for col in df.columns if col not in predefined_cols]
    chunk_size = 10
    total_chunks = ceil(len(tag_columns) / chunk_size)
    current_row = 1

    try:
        from footer import extract_template_with_placeholders, insert_footer_into_worksheet
        footer_template = extract_template_with_placeholders("template.xlsx", "A2:O6")
    except ImportError:
        footer_template = {"template": []}
        print("Warning: Footer module not found. Footer insertion will be skipped.")

    for i in range(total_chunks):
        footer_values = {
            "division": {"value": "PRAYAGRAJ Division", "size": 22},
            "railways": {"value": "NC RAILWAYS", "size": 22},
            "station-name": {"value": "MALWAN(MWH)", "size": 22},
            "station-id": {"value": "Section Id: 37111", "size": 22},
            "date": {"value": "25-07-2025", "size": 22},
            "total-pages": {"value": str(total_chunks), "size": 22},
            "current-page": {"value": str(i + 1), "size": 22}
        }
        start = i * chunk_size
        end = start + chunk_size
        tag_chunk = tag_columns[start:end]
        temp_df = df[predefined_cols + tag_chunk]

        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=15)
        header_cell = ws.cell(row=current_row, column=1, value=tag_title)
        header_cell.font = Font(name='Times New Roman', size=14, bold=True)
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        header_cell.border = border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        for r_idx, row in enumerate(dataframe_to_rows(temp_df, index=False, header=True), start=current_row):
            for c_idx, value in enumerate(row, start=1):
                actual_col = c_idx + 2 if c_idx > 1 else c_idx
                if c_idx == 1:
                    ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=3)
                    ws.cell(row=r_idx, column=1, value=value)
                    for col in range(1, 4):
                        ws.cell(row=r_idx, column=col).border = border
                else:
                    ws.cell(row=r_idx, column=actual_col, value=value)

                cell = ws.cell(row=r_idx, column=actual_col if c_idx > 1 else 1)
                is_header = r_idx == current_row
                is_predefined = c_idx <= 3
                cell.font = bold_font if (is_predefined and not is_header) else (bold_font if is_header else font)
                cell.alignment = align
                cell.border = border
                if is_header and c_idx <= 3:
                    cell.fill = fill

        for col_letter in ['A', 'B', 'C']:
            ws.column_dimensions[col_letter].width = 20
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 35

        num_tag_cols = 10
        min_width = 20
        for col_offset in range(num_tag_cols):
            col = 6 + col_offset
            col_letter = get_column_letter(col)
            max_length = 0
            if col_offset < len(temp_df.columns) - 3:
                for row_offset in range(temp_df.shape[0] + 1):
                    cell = ws.cell(row=current_row + row_offset, column=col)
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            width = max(min_width, min(max_length + 4, 40))
            ws.column_dimensions[col_letter].width = width

        for row_offset in range(temp_df.shape[0] + 1):
            r = current_row + row_offset
            max_lines = 1
            for col in range(1, 4 + len(temp_df.columns) - 1):
                cell = ws.cell(row=r, column=col)
                if cell.value:
                    text = str(cell.value)
                    line_breaks = text.count('\n')
                    wrapped_lines = len(text) // 40 + 1
                    total_lines = max(line_breaks + 1, wrapped_lines)
                    max_lines = max(max_lines, total_lines)
            ws.row_dimensions[r].height = max_lines * 15 + 20 if max_lines > 2 else 40

        current_row += temp_df.shape[0] + 4
        try:
            insert_footer_into_worksheet(ws, footer_template, footer_values, current_row)
        except NameError:
            print(f"Footer skipped for chunk {i+1} in sheet with title: {tag_title}")
        current_row += len(footer_template["template"]) + 6



def process_input_sheet(sheet_name, df):
    print(f"\nProcessing sheet: {sheet_name}")

    df = df.dropna(how='all').reset_index(drop=True)
    if df.empty:
        print(f"Sheet {sheet_name} is empty after dropping NA rows.")
        return None

    df.columns = [str(col).strip() for col in df.columns]
    
    df['original_index'] = df[df.columns[0]]
    if df['original_index'].duplicated().any():
        print(f"Duplicate index labels found in {sheet_name}: {df['original_index'][df['original_index'].duplicated()].tolist()}")
        df['original_index'] = df['original_index'].astype(str) + '_' + df.groupby('original_index').cumcount().astype(str)
        df['original_index'] = df['original_index'].str.replace('_0$', '', regex=True)

    df.set_index('original_index', inplace=True)
    df.index = df.index.astype(str).str.strip()

    tag_cols = extract_tag_columns(df.columns)
    print(f"Tag columns found: {tag_cols}")
    if not tag_cols:
        print(f"No valid tag columns in sheet {sheet_name}. Regex used: r'^\\d+/[MD]$'")
        return None

    field_desc = list(df.index)
    bit_positions = df.loc[:, "BIT POSITION"] if "BIT POSITION" in df.columns else [""] * len(field_desc)
    sizes = df.loc[:, "Size (Bits)"] if "Size (Bits)" in df.columns else [""] * len(field_desc)

    output_df = pd.DataFrame({
        "FIELD NAME / DESCRIPTION": field_desc + ["CRC", "PAGE -X", "PAGE -Y"],
        "BIT POSITION": list(bit_positions) + ["Y63 - Y34", "", ""],
        "Size (Bits)": list(sizes) + [30, 64, 64]
    })

    nt_field_mapping = {
        'uctypeofTag': 'Type of Tag (9- Normal, 10 - LC, 11- Adj.Line ,12-Junction)',
        'uc_version': 'Version(As per Spec 4.0)',
        'uiUniqueID': 'Unique ID of RFID Tag Set',
        'fAbsLoc': 'Absolute Loc In  meters',
        'ucTagPlacement': 'Tag Placement(0-InL, 1- SIG(N), 2-SIG(R), 3-Tout, 4-Exit(N),5-Exit(R), 6-SIG(N/R),7-exit tag both directions, 8-Dead stopin Nominal,9- Dead stop in Reverse)',
        'AbsoluteLocationReset': 'Tag Duplication (0-Main tag,1-Dup tag)',
        'nominal': {
            'ucTin': 'TIN in Nominal Direction',
            'StationId': 'Station ID in Nominal Direction',
            'comMark': 'Communication in Nominal Direction  (0- Required, 1- Not Required).',
            'secType': 'Section type in Nominal Direction ( 0-Station, 1- Abs Blk, 2-Auto, 3- VBlk)'
        },
        'reverse': {
            'ucTin': 'TIN in Reverse Direction',
            'StationId': 'Station ID in Reverse Direction',
            'comMark': 'Communication in Reverse Direction. (0- Required, 1- Not Required).',
            'secType': 'Section type in Reverse Direction ( 0-Station, 1- Abs Blk, 2-Auto, 3- VBlk)'
        }
    }

    aline_field_mapping = {
        'uctypeofTag': 'Type of Tag (9- Normal, 10 - LC, 11- Adj.Line ,12-Junction)',
        'uc_version': 'Version(As per Spec 4.0)',
        'uiUniqueID': 'Unique ID of RFID Tag Set',
        'fAbsLoc': 'Absolute Loc In  meters',
        'nominal': {
            'ucTin': 'TIN in Nominal Direction'
        },
        'reverse': {
            'ucTin': 'TIN in Reverse Direction'
        },
        'AdjLine1_tin': 'Adjacent Line-1 TIN',
        'AdjLine2_tin': 'Adjacent Line-2 TIN',
        'AdjLine3_tin': 'Adjacent Line-3 TIN',
        'AdjLine4_tin': 'Adjacent Line-4 TIN',
        'AdjLine5_tin': 'Adjacent Line-5 TIN',
        'ucTagDuplication': 'Tag Duplication (0-Main tag,1-Dup tag)'
    }

    adj_field_mapping = {
        'uctypeofTag': 'Type of Tag (9- Normal, 10 - LC, 11- Adj.Line ,12-Junction)',
        'uc_version': 'Version(As per Spec 4.0)',
        'uiUniqueID': 'Unique ID of RFID Tag Set',
        'fAbsLoc1': 'Absolute Loc In  meters-1',
        'fAbsLoc2': 'Absolute Loc In  meters-2',
        'nominal': {
            'ucTin': 'TIN 1'
        },
        'reverse': {
            'ucTin': 'TIN 2'
        },
        'dirResetAbsLoc1': 'Direction reset absolute location-1',
        'dirResetAbsLoc2': 'Direction reset absolute location-2',
        'locCorrectionType': 'Location Correction Type',
        'reserved': 'Reserved',
        'secTypeNominal': 'Section type in nominal direction',
        'secTypeReverse': 'Section type in reverse direction',
        'reserved1': 'Reserved_1',
        'tagType': 'Tag type',
        'comMarkNominal': 'Communication in nominal direction',
        'comMarkReverse': 'Communication in reverse direction'
    }

    tag_data = {}
    for col in tag_cols:
        values = []
        try:
            for row in field_desc:
                try:
                    val = df.at[row, col]
                    if isinstance(val, pd.Series):
                        val = val.iloc[0]
                    values.append(val)
                except (KeyError, IndexError):
                    values.append("")

            if sheet_name == 'NT':
                uctypeofTag = int(df.at[nt_field_mapping['uctypeofTag'], col]) if nt_field_mapping['uctypeofTag'] in df.index else 9
                uc_version = int(df.at[nt_field_mapping['uc_version'], col]) if nt_field_mapping['uc_version'] in df.index else 1
                uiUniqueID = int(df.at[nt_field_mapping['uiUniqueID'], col]) if nt_field_mapping['uiUniqueID'] in df.index else 0
                fAbsLoc = int(df.at[nt_field_mapping['fAbsLoc'], col]) if nt_field_mapping['fAbsLoc'] in df.index and df.at[nt_field_mapping['fAbsLoc'], col] != 'UNKNOWN' else 0
                ucTagPlacement = int(df.at[nt_field_mapping['ucTagPlacement'], col]) if nt_field_mapping['ucTagPlacement'] in df.index else 0
                AbsoluteLocationReset = int(df.at[nt_field_mapping['AbsoluteLocationReset'], col]) if nt_field_mapping['AbsoluteLocationReset'] in df.index else 0

                nominal_ucTin = int(df.at[nt_field_mapping['nominal']['ucTin'], col]) if nt_field_mapping['nominal']['ucTin'] in df.index else 0
                nominal_StationId = int(df.at[nt_field_mapping['nominal']['StationId'], col]) if nt_field_mapping['nominal']['StationId'] in df.index else 0
                nominal_comMark = int(df.at[nt_field_mapping['nominal']['comMark'], col]) if nt_field_mapping['nominal']['comMark'] in df.index else 0
                nominal_secType = int(df.at[nt_field_mapping['nominal']['secType'], col]) if nt_field_mapping['nominal']['secType'] in df.index else 0

                reverse_ucTin = int(df.at[nt_field_mapping['reverse']['ucTin'], col]) if nt_field_mapping['reverse']['ucTin'] in df.index else 0
                reverse_StationId = int(df.at[nt_field_mapping['reverse']['StationId'], col]) if nt_field_mapping['reverse']['StationId'] in df.index else 0
                reverse_comMark = int(df.at[nt_field_mapping['reverse']['comMark'], col]) if nt_field_mapping['reverse']['comMark'] in df.index else 0
                reverse_secType = int(df.at[nt_field_mapping['reverse']['secType'], col]) if nt_field_mapping['reverse']['secType'] in df.index else 0

                tag = TagInfoNT(
                    uctypeofTag=uctypeofTag,
                    uc_version=uc_version,
                    uiUniqueID=uiUniqueID,
                    fAbsLoc=fAbsLoc,
                    ucTagPlacement=ucTagPlacement,
                    AbsoluteLocationReset=AbsoluteLocationReset,
                    stDir=[
                        TagDirNT(nominal_ucTin, nominal_StationId, nominal_comMark, nominal_secType),
                        TagDirNT(reverse_ucTin, reverse_StationId, reverse_comMark, reverse_secType)
                    ]
                )

                page1, page2, crc = calculate_values_nt(tag)
                def format_page(page: bytearray) -> str:
                    trimmed = page[:8]
                    while trimmed and trimmed[-1] == 0x00 and len(trimmed) > 6:
                        trimmed = trimmed[:-1]
                    return ''.join(f"{b:02x}" for b in trimmed)

                # Format output
                crc_str = f"{crc:08x}"
                page_x = format_page(page1)
                page_y = format_page(page2)


            elif sheet_name == 'AlineT':
                uctypeofTag = int(df.at[aline_field_mapping['uctypeofTag'], col]) if aline_field_mapping['uctypeofTag'] in df.index else 11
                uc_version = int(df.at[aline_field_mapping['uc_version'], col]) if aline_field_mapping['uc_version'] in df.index else 1
                uiUniqueID = int(df.at[aline_field_mapping['uiUniqueID'], col]) if aline_field_mapping['uiUniqueID'] in df.index else 0
                fAbsLoc = int(df.at[aline_field_mapping['fAbsLoc'], col]) if aline_field_mapping['fAbsLoc'] in df.index and df.at[aline_field_mapping['fAbsLoc'], col] != 'UNKNOWN' else 0
                nominal_ucTin = int(df.at[aline_field_mapping['nominal']['ucTin'], col]) if aline_field_mapping['nominal']['ucTin'] in df.index else 0
                reverse_ucTin = int(df.at[aline_field_mapping['reverse']['ucTin'], col]) if aline_field_mapping['reverse']['ucTin'] in df.index else 0
                AdjLine1_tin = int(df.at[aline_field_mapping['AdjLine1_tin'], col]) if aline_field_mapping['AdjLine1_tin'] in df.index else 0
                AdjLine2_tin = int(df.at[aline_field_mapping['AdjLine2_tin'], col]) if aline_field_mapping['AdjLine2_tin'] in df.index else 0
                AdjLine3_tin = int(df.at[aline_field_mapping['AdjLine3_tin'], col]) if aline_field_mapping['AdjLine3_tin'] in df.index else 0
                AdjLine4_tin = int(df.at[aline_field_mapping['AdjLine4_tin'], col]) if aline_field_mapping['AdjLine4_tin'] in df.index else 0
                AdjLine5_tin = int(df.at[aline_field_mapping['AdjLine5_tin'], col]) if aline_field_mapping['AdjLine5_tin'] in df.index else 0
                ucTagDuplication = int(df.at[aline_field_mapping['ucTagDuplication'], col]) if aline_field_mapping['ucTagDuplication'] in df.index else 0
                
                tag = TagInfoAline(
                    uctypeofTag=uctypeofTag,
                    uc_version=uc_version,
                    uiUniqueID=uiUniqueID,
                    fAbsLoc=fAbsLoc,
                    stDir=[
                        TagDirAline(nominal_ucTin),
                        TagDirAline(reverse_ucTin)
                    ],
                    AdjLine1_tin=AdjLine1_tin,
                    AdjLine2_tin=AdjLine2_tin,
                    AdjLine3_tin=AdjLine3_tin,
                    AdjLine4_tin=AdjLine4_tin,
                    AdjLine5_tin=AdjLine5_tin,
                    ucTagDuplication=ucTagDuplication
                )

                result = calculate_values_aline(tag)
                crc_str = f"{result.crc:08X}".lower()
                page_x = result.page_x.hex().lower()
                page_y = result.page_y.hex().lower()

            elif sheet_name == 'AdjT':
                uctypeofTag = int(df.at[adj_field_mapping['uctypeofTag'], col]) if adj_field_mapping['uctypeofTag'] in df.index else 12
                uc_version = int(df.at[adj_field_mapping['uc_version'], col]) if adj_field_mapping['uc_version'] in df.index else 1
                uiUniqueID = int(df.at[adj_field_mapping['uiUniqueID'], col]) if adj_field_mapping['uiUniqueID'] in df.index else 0
                fAbsLoc1 = int(df.at[adj_field_mapping['fAbsLoc1'], col]) if adj_field_mapping['fAbsLoc1'] in df.index and df.at[adj_field_mapping['fAbsLoc1'], col] != 'UNKNOWN' else 0
                fAbsLoc2 = int(df.at[adj_field_mapping['fAbsLoc2'], col]) if adj_field_mapping['fAbsLoc2'] in df.index and df.at[adj_field_mapping['fAbsLoc2'], col] != 'UNKNOWN' else 0
                nominal_ucTin = int(df.at[adj_field_mapping['nominal']['ucTin'], col]) if adj_field_mapping['nominal']['ucTin'] in df.index else 0
                reverse_ucTin = int(df.at[adj_field_mapping['reverse']['ucTin'], col]) if adj_field_mapping['reverse']['ucTin'] in df.index else 0
                dirResetAbsLoc1 = int(df.at[adj_field_mapping['dirResetAbsLoc1'], col]) if adj_field_mapping['dirResetAbsLoc1'] in df.index else 0
                dirResetAbsLoc2 = int(df.at[adj_field_mapping['dirResetAbsLoc2'], col]) if adj_field_mapping['dirResetAbsLoc2'] in df.index else 0
                locCorrectionType = int(df.at[adj_field_mapping['locCorrectionType'], col]) if adj_field_mapping['locCorrectionType'] in df.index else 0
                reserved = int(df.at[adj_field_mapping['reserved'], col]) if adj_field_mapping['reserved'] in df.index else 0
                secTypeNominal = int(df.at[adj_field_mapping['secTypeNominal'], col]) if adj_field_mapping['secTypeNominal'] in df.index else 0
                secTypeReverse = int(df.at[adj_field_mapping['secTypeReverse'], col]) if adj_field_mapping['secTypeReverse'] in df.index else 0
                reserved1 = int(df.at[adj_field_mapping['reserved1'], col]) if adj_field_mapping['reserved1'] in df.index else 0
                tagType = int(df.at[adj_field_mapping['tagType'], col]) if adj_field_mapping['tagType'] in df.index else 0
                comMarkNominal = int(df.at[adj_field_mapping['comMarkNominal'], col]) if adj_field_mapping['comMarkNominal'] in df.index else 0
                comMarkReverse = int(df.at[adj_field_mapping['comMarkReverse'], col]) if adj_field_mapping['comMarkReverse'] in df.index else 0


                tag = TagInfoAdj(
                    uctypeofTag=uctypeofTag,
                    uc_version=uc_version,
                    uiUniqueID=uiUniqueID,
                    stDir=[
                        TagDirAdj(nominal_ucTin, secTypeNominal, dirResetAbsLoc1, comMarkNominal, fAbsLoc1),
                        TagDirAdj(reverse_ucTin, secTypeReverse, dirResetAbsLoc2, comMarkReverse, fAbsLoc2)
                    ],
                    locCorrectionType=locCorrectionType,
                    reserved=reserved,
                    reserved1=reserved1,
                    tagType=tagType
                )
                
                page1, page2, crc = calculate_values_adj(tag)
                def format_page(page: bytearray) -> str:
                    trimmed = page[:8]
                    while trimmed and trimmed[-1] == 0x00 and len(trimmed) > 6:
                        trimmed = trimmed[:-1]
                    return ''.join(f"{b:02x}" for b in trimmed)

                # Format output
                crc_str = f"{crc:08x}"
                page_x = format_page(page1)
                page_y = format_page(page2)

            else:
                crc_str = 0
                page_x = '0 0 0 0 0 0 0 0'
                page_y = '0 0 0 0 0 0 0 0'

            values += [crc_str, page_x, page_y]
            tag_data[str(col)] = values

        except Exception as e:
            print(f"Error processing column {col} in sheet {sheet_name}: {e}")
            values = [""] * len(field_desc) + [0, '0 0 0 0 0 0 0 0', '0 0 0 0 0 0 0 0']
            tag_data[str(col)] = values

    if tag_data:
        tag_df = pd.DataFrame(tag_data, index=output_df.index)
        output_df = pd.concat([output_df, tag_df], axis=1)

    return output_df

def process_excel():
    try:
        if 'input_path' not in request.form or 'output_path' not in request.form:
            return jsonify({"error": "Missing input or output path"}), 400

        input_path = request.form['input_path']
        output_path = request.form['output_path']
        
        if not input_path.endswith('.xlsx'):
            return jsonify({"error": "Invalid input file format. Please provide an .xlsx file path"}), 400

        if not os.path.isfile(input_path):
            return jsonify({"error": "Input file does not exist"}), 400

        if not os.path.isdir(os.path.dirname(output_path)):
            return jsonify({"error": "Invalid output directory"}), 400
        
        drive = os.path.splitdrive(output_path)[0]
        if drive and not os.path.exists(drive):
            return jsonify({"error": f"Drive {drive} does not exist or is not accessible"}), 400

        # Create full output directory if not present
        if not os.path.exists(output_path):
            os.makedirs(output_path)

        # Generate output file name by appending '_formatted' to input file name
        input_filename = os.path.splitext(os.path.basename(input_path))[0]
        base_output_filename = f"{input_filename}_formatted"
        output_extension = '.xlsx'
        output_file = os.path.join(output_path, f"{base_output_filename}{output_extension}")
        
        # Check for existing files and append (1), (2), etc., if necessary
        if os.path.isfile(output_file):
            os.remove(output_file)
        
        wb = Workbook()
        all_sheets = pd.read_excel(input_path, sheet_name=None, header=1)
        all_raw = pd.read_excel(input_path, sheet_name=None, header=None)
        first = True
        processed_sheets = 0

        for sheet_name, df in all_sheets.items():
            raw_df = all_raw[sheet_name]
            tag_title = str(raw_df.iloc[0, 0]) if not raw_df.empty else "TAG"

            formatted_df = process_input_sheet(sheet_name, df)
            if formatted_df is None:
                print(f"Skipping sheet {sheet_name} — no valid tag data.")
                continue

            ws = wb.active if first else wb.create_sheet()
            ws.title = sheet_name
            first = False
            processed_sheets += 1

            format_sheet(ws, formatted_df, tag_title)

        if processed_sheets == 0:
            return jsonify({"error": "No sheets processed. Output file not saved."}), 400

        wb.save(output_file)
        print(f"✅ Final Excel saved: {output_file}")

        return send_file(
            output_file,
            as_attachment=True,
            download_name=os.path.basename(output_file),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return jsonify({"error": f"Processing failed: {str(e)}"}), 500


