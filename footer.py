import copy
import re
from openpyxl import load_workbook, Workbook
from openpyxl.utils import range_boundaries
from openpyxl.utils.cell import range_boundaries
from openpyxl.styles import Alignment, Font


def extract_template_with_placeholders(file_path, range_str):
    """
    Extracts a formatted range from Excel and stores as a reusable template with placeholder support.
    """
    wb = load_workbook(file_path)
    ws = wb.active

    min_col, min_row, max_col, max_row = range_boundaries(range_str)

    template_cells = []
    merged_ranges = []

    for r in range(min_row, max_row + 1):
        row = []
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            row.append({
                "value": str(cell.value) if cell.value is not None else "",
                "font": copy.copy(cell.font),
                "border": copy.copy(cell.border),
                "fill": copy.copy(cell.fill),
                "alignment": copy.copy(cell.alignment),
            })
        template_cells.append(row)

    for merged in ws.merged_cells.ranges:
        mc_min_col, mc_min_row, mc_max_col, mc_max_row = range_boundaries(str(merged))
        if min_row <= mc_min_row <= max_row:
            merged_ranges.append(str(merged))

    return {
        "template": template_cells,
        "start_row": min_row,
        "start_col": min_col,
        "merged_cells": merged_ranges
    }


def extract_placeholders(template_data):
    """
    Scans the template for all {{placeholders}} and returns a unique list of field names.
    """
    placeholders = set()
    for row in template_data["template"]:
        for cell in row:
            matches = re.findall(r"{{(.*?)}}", cell["value"])
            placeholders.update(matches)
    return list(placeholders)


def generate_excel_from_template(template_data, output_path, variables: dict):
    """
    Generates an Excel file using the template and replaces placeholders with provided data.
    Leaves other non-placeholder content unchanged.
    """
    wb = Workbook()
    ws = wb.active

    start_row = template_data["start_row"]
    start_col = template_data["start_col"]
    template = template_data["template"]

    for r_idx, row in enumerate(template):
        for c_idx, cell_data in enumerate(row):
            cell = ws.cell(row=start_row + r_idx, column=start_col + c_idx)

            # Replace placeholders, leave rest of the text as-is
            value = cell_data["value"]
            for key, val in variables.items():
                value = value.replace(f"{{{{{key}}}}}", str(val))

            cell.value = value
            cell.font = cell_data["font"]
            cell.fill = cell_data["fill"]
            cell.border = cell_data["border"]
            cell.alignment = cell_data["alignment"]

    for merge in template_data["merged_cells"]:
        ws.merge_cells(merge)

    wb.save(output_path)
    print(f"✅ File saved: {output_path}")


def insert_footer_into_worksheet(ws, template_data, variables, start_row):
    """
    Inserts the footer into an existing worksheet (ws) at the specified start_row.
    Ensures that all text, including merged cells, is center-aligned and each row is at least 30 height.
    Supports custom font sizes for dynamic values using {"value": ..., "size": ...}
    """
    start_col = template_data["start_col"]
    template = template_data["template"]
    original_start_row = template_data["start_row"]

    for r_idx, row in enumerate(template):
        for c_idx, cell_data in enumerate(row):
            row_num = start_row + r_idx
            col_num = start_col + c_idx
            cell = ws.cell(row=row_num, column=col_num)

            # Get original template value
            original_text = cell_data["value"]
            final_text = original_text
            applied_font_size = 11  # default

            # Replace placeholders with values
            for key, val in variables.items():
                placeholder = f"{{{{{key}}}}}"
                if placeholder in final_text:
                    # Support for {"value": ..., "size": ...}
                    if isinstance(val, dict):
                        replacement_text = str(val.get("value", ""))
                        font_size = val.get("size", 11)
                    else:
                        replacement_text = str(val)
                        font_size = 11

                    final_text = final_text.replace(placeholder, replacement_text)

                    # If entire cell is just this placeholder, use its font size
                    if original_text.strip() == placeholder:
                        applied_font_size = font_size

            cell.value = final_text

            # Apply original font settings with modified font size
            original_font = cell_data["font"]
            cell.font = Font(
                name=original_font.name,
                bold=original_font.bold,
                italic=original_font.italic,
                vertAlign=original_font.vertAlign,
                underline=original_font.underline,
                strike=original_font.strike,
                color=original_font.color,
                size=applied_font_size  # 👈 custom font size
            )

            # Apply other styles
            cell.fill = cell_data["fill"]
            cell.border = cell_data["border"]
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Adjust row height
        ws.row_dimensions[start_row + r_idx].height = max(ws.row_dimensions[start_row + r_idx].height or 0, 50)

    # Recreate merged cells with adjusted positions
    for merge in template_data["merged_cells"]:
        min_col, min_row, max_col, max_row = range_boundaries(merge)
        row_offset = start_row - original_start_row

        new_merge_range = f"{ws.cell(row=min_row + row_offset, column=min_col).coordinate}:{ws.cell(row=max_row + row_offset, column=max_col).coordinate}"
        ws.merge_cells(new_merge_range)

        # Set alignment for merged cell
        merged_top_left_cell = ws.cell(row=min_row + row_offset, column=min_col)
        merged_top_left_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
